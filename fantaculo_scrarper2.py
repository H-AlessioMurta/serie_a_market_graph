import pandas as pd
import numpy as np
import requests
import openpyxl
import matplotlib.pyplot as plt
from datetime import datetime
from bs4 import BeautifulSoup
from openpyxl.drawing.image import Image
import io
import unicodedata

id_collection=pd.DataFrame(columns=['nome','id_fantacalcio','id_transfermarket','id_fantalgoritmo'])
tabella=pd.DataFrame(columns = [
    'Nome Giocatore',           # giocatore['name']
    'Squadra',                  # giocatore['team']
    'Ruolo',                    # giocatore['role']
    'Slot',                     # giocatore['features']['slot']
    'PMA',                      # giocatore['features']['pma']
    'PFC',                      # giocatore['features']['pfc']
    'FVM Classic 350',          # (int(giocatore['FVM']['Classic'])/1000)*350
    'Fantamedia Attesa',        # giocatore['features']['expectedFantamedia']
    'Titolarità Attesa',        # giocatore['features']['expectedTitolarita']
    'Probabilità Rigore',       # giocatore['features']['penaltyProbability']
    'Probabilità Punizione',    # giocatore['features']['freeKickProbability']
    'Range PMA',                # giocatore['features']['pmaRange']
    'Range PFC',                # giocatore['features']['pfcRange']
    'Voto Base Ultimi 3 Anni',  # giocatore['features']['lastThreeYearVotoBase']
    'Fantamedia Ultimi 3 Anni', # giocatore['features']['lastThreeYearFantamedia']
    'Titolarità Ultimi 3 Anni', # giocatore['features']['lastThreeYearTitolarity']
    'Voto Base Ultimi 5 Anni',  # giocatore['features']['lastFiveYearVotoBase']
    'Fantamedia Ultimi 5 Anni', # giocatore['features']['lastFiveYearFantamedia']
    'Titolarità Ultimi 5 Anni', # giocatore['features']['lastFiveYearTitolarity']
    'Voto Base Anno Scorso',    # giocatore['features']['lastYearVotoBase']
    'Fantamedia Anno Scorso',   # giocatore['features']['lastYearFantamedia']
    'Titolarità Anno Scorso',   # giocatore['features']['lastYearTitolarity']
    'Voto Base Stagione Corrente', # giocatore['features']['currentSeasonVotoBase']
    'Fantamedia Stagione Corrente', # giocatore['features']['currentSeasonFantamedia']
    'Titolarità Stagione Corrente', # giocatore['features']['currentSeasonTitolarity']
    'Piede',
    'Altezza',                  # giocatore['Altezza']
    'City',
    'Data di Nascita',          # giocatore['Nato il']
    'Nazionalità',              # giocatore['Nazionalità']
    'Valore Corrente',          # giocatore['current']
    'Valore Massimo',           # giocatore['highest']
    'Età',                      # giocatore['age']
    'Descrizione',               # giocatore['descrizione'],
    'Fine Contratto',
    'Posizione'
])
grafici=pd.DataFrame(columns=['Giocatore','Grafico'])
storico_performances = pd.DataFrame(columns=['Giocatore', 'Anno', 'Competition', 'Club', 'Appearances','Goals', 'Assists','GoalParticipationPercentage','SubstitutionsOn','SubstitutionsOff', 'YellowCards', 'RedCards', 'MinutesPlayed','minutesPerGoal'])
storico_infortuni=pd.DataFrame(columns=['Giocatore','Anno','totalDays','totalInjuryRate','totalGamesMissed'])
#body_listone={"credits": 350,"flagNoGoal": False,"flagModDefense": True,"competitionParticipants": 12,"player_name": None,"flagSvincolati": False,"flagMantra": False,"auctionType": "pma","flag_clean_sheet": False, }   

transfermarket_teams = {
    'Inter': 46,
    'Milan': 5,
    'Juventus': 506,
    'Atalanta': 800,
    'Bologna': 1025,
    'Roma': 12,
    'Lazio': 398,
    'Fiorentina': 430,
    'Torino': 416,
    'Napoli': 6195,
    'Genoa': 252,
    'Monza': 2919,
    'Verona': 276,
    'Lecce': 1005,
    'Udinese': 410,
    'Cagliari': 1390,
    'Empoli': 749,
    'Parma': 130,
    'Como': 1047,
    'Venezia': 607
}

def transfermarket_infortuni():
    headers = { "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"}
    html = requests.get("https://www.transfermarkt.it/serie-a/verletztespieler/wettbewerb/IT1", headers=headers)
    soup = BeautifulSoup(html.content, "html.parser")
    rows = soup.find_all('tr', class_=['odd', 'even'])
    lista_infortunati=[]
    for row in rows:
        player_name = row.find('a', title=True).text
        injury_type = row.find_all('td')[5].text.strip()
        data_rientro = row.find_all('td')[6].text.strip()
        market_value = row.find_all('td')[-1].text.strip()
        lista_infortunati.append({
            "Giocatore": player_name,
            "Infortunio": injury_type,
            "Data rientro": data_rientro,
            "Valore di Mercato": market_value
        })
    return lista_infortunati


def get_hometown(link):
    hometown_url=f"https://www.transfermarkt.it{link}"
    headers = { "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"}
    home_player = requests.get(hometown_url, headers=headers)
    soup = BeautifulSoup(home_player.content, "html.parser")
    li_tag = soup.find_all('li', class_='data-header__label')
# Estrarre il testo all'interno di <span> con itemprop="birthPlace"
    birth_place = "Ignoto figlio di mignotta"
    for x in li_tag:
        if x.find('span', itemprop='birthPlace'):
            birth_place = birth_place = x.find('span', itemprop='birthPlace').get_text(strip=True)
    return birth_place



def transfermarket_infos(mapping_id,team,nome):
    ecezzioni_note={
            "sangare": 962125,
            "nava": 815563,
            "liberali": 988964,
            "adzic": 944570,
            "mbangula": 654991,
            "ekhator": 934878,
            "marin": 1041614,
            "bellanova": 357992,
            "kalolu": 585949,
            "nicolussi": 430280,
            "zambo":354361,
            "vos": 738476,
            "adli":395236
        }
    id=-1
    link=''
    if len(nome.split(" ")[0]) >2:
        nome_da_cercare=nome.split(" ")[0]
    else:
        nome_da_cercare=nome.split(" ")[0]+' '+nome.split(" ")[1]
    nome_da_cercare=nome_da_cercare.lower().replace("'","")
    for a_player in mapping_id[team]:
         if nome_da_cercare in a_player['name'].lower():
            id=a_player['id']
         if nome_da_cercare == "dambrosio":
            id=55769
         if nome_da_cercare == "yildiz":
            id=845654
    if nome == "MARTINEZ L.":
        id=406625
    if nome == "CAMARDA":
        id=1058368
    if nome_da_cercare in ecezzioni_note:
        id=ecezzioni_note[nome_da_cercare]
    if id == -1:
       print(f"non trovato { nome_da_cercare } nel {team} che id transfermarket uso? ")
       id=input("Inserisci un valore: ")
    for a_player in mapping_id [team]:
        if a_player['id']==id:
            link=a_player['link']
    if id != -1:
       tm_datas = {}
       #print(f"Trovato id di {nome} con valore {id}")
       performance_url=f"https://www.transfermarkt.it/ceapi/player/{id}/performance"
       performance_result=transfermarket_json(performance_url)
       for x in performance_result:
           if x['nameSeason'] != '24/25' and x['nameSeason']  != None :
                tm_datas['possibleGames_' + x['nameSeason']] = x['possibleGames']
                tm_datas['gamesPlayed_' + x['nameSeason']] = x['gamesPlayed']
                tm_datas['goalsScored_' + x['nameSeason']] = x['goalsScored']
                tm_datas['assists_' + x['nameSeason']] = x['assists']
                tm_datas['yellowCards_' + x['nameSeason']] = x['yellowCards']
                tm_datas['secondYellowCards_' + x['nameSeason']] = x['secondYellowCards']
                tm_datas['redCards_' + x['nameSeason']] = x['redCards']
                tm_datas['startElevenPercent_' + x['nameSeason']] = x['startElevenPercent']
                tm_datas['minutesPlayedPercent_' + x['nameSeason']] = x['minutesPlayedPercent']
                tm_datas['goalsContributedPercent_' + x['nameSeason']] = x['goalsContributedPercent']
                tm_datas['concededGoals_' + x['nameSeason']] = x['concededGoals']
                tm_datas['cleanSheets_' + x['nameSeason']] = x['cleanSheets']
                tm_datas['blockedPenaltyPercent_' + x['nameSeason']] = x['blockedPenaltyPercent']
                tm_datas['minutesPlayed_' + x['nameSeason']] = x['minutesPlayed']
       market_values_url=f'https://www.transfermarkt.it/ceapi/marketValueDevelopment/graph/{id}'
       market_values_result=transfermarket_json(market_values_url)
       #print(market_values_result['list'])
       popped_marked=[]
       actual_age=0
       for x in market_values_result['list']:
           x.pop('wappen')
           popped_marked.append(x)
           if actual_age<int(x['age']):
               actual_age=int(x['age'])
       tm_datas['market_values']=popped_marked
       tm_datas['current'] = market_values_result['current']
       tm_datas['highest'] = market_values_result['highest']
       tm_datas['age']=actual_age
       tm_datas['city']=get_hometown(link)
       tm_datas['id_transfermarket']=id
       return tm_datas

def transfermarket_json(tmurl):
    headers = { "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"}
    x = requests.get(tmurl, headers=headers)
    return x.json()

def aggiorna_dati_fantaculo():
    print('Chiamata listone aggiornato - Fantaculo')
    
    x = requests.get('https://fantaculo.it/leghe-srv/api/v1/aste/listone?credits=350&flagNoGoal=false&flagModDefense=true&competitionParticipants=12&name=null&flagSvincolati=false&flagMantra=false&auctionType=pma')
    return x.json()

def get_by_name(name,id,mapping_id):
        player_url=f'https://fantaculo.it/leghe-srv/api/v1/aste/player?credits=350&flagNoGoal=false&flagModDefense=true&competitionParticipants=12&name={name}&flagSvincolati=false&flagMantra=false&auctionType=pma'
        x = requests.get(player_url)
        giocatore=x.json()
        print("inizo chiamate")
        giocatore.update(transfermarket_infos(mapping_id,giocatore['team'],giocatore['name']))
        print("----fine chiamata transfermarket")
        giocatore.update(fantacalcio_calciatore(giocatore['team'].lower(),giocatore['name'].lower(),id))
        print("----fine chiamata fantacalcio.it")
        giocatore.update(fantaalgoritmo(giocatore['name'].lower(),id))
        print("----fine chiamata fantaalgortimo")
        #print(giocatore
        return giocatore

def scraper(mapping_id,lista_infortuni):
    df_lista_infortuni= pd.DataFrame(lista_infortuni)
    listone=aggiorna_dati_fantaculo()
    conta = 0
    for i in listone:
        conta+=1
        print(f'N {conta} chiamata per {i["name"]}')
        name=i["name"]
        id=i['idFantacalcio']
        try:
            if name != 'RUGANI':
                giocatore=get_by_name(name,id,mapping_id)
                nuovo_rigo = [
                    giocatore['name'],
                    giocatore['team'],
                    giocatore['role'],
                    giocatore['features']['slot'],
                    giocatore['features']['pma'],
                    giocatore['features']['pfc'],
                    (int(giocatore['FVM']['Classic'])/1000)*350,
                    giocatore['features']['expectedFantamedia'],
                    giocatore['features']['expectedTitolarita'],
                    giocatore['features']['penaltyProbability'],
                    giocatore['features']['freeKickProbability'],
                    giocatore['features']['pmaRange'],
                    giocatore['features']['pfcRange'],
                    giocatore['features']['lastThreeYearVotoBase'],
                    giocatore['features']['lastThreeYearFantamedia'],
                    giocatore['features']['lastThreeYearTitolarity'],
                    giocatore['features']['lastFiveYearVotoBase'],
                    giocatore['features']['lastFiveYearFantamedia'],
                    giocatore['features']['lastFiveYearTitolarity'],
                    giocatore['features']['lastYearVotoBase'],
                    giocatore['features']['lastYearFantamedia'],
                    giocatore['features']['lastYearTitolarity'],
                    giocatore['features']['currentSeasonVotoBase'],
                    giocatore['features']['currentSeasonFantamedia'],
                    giocatore['features']['currentSeasonTitolarity'],
                    giocatore['Piede'],
                    giocatore['Altezza'],
                    giocatore['city'],
                    giocatore['Nato il'],
                    giocatore['Nazionalità'],
                    giocatore['current'],
                    giocatore['highest'],
                    giocatore['age'],
                    giocatore['descrizione'],
                    giocatore['fine_contratto'],
                    giocatore['Posizione']
                ]
                if 'injuries' in giocatore and  'performances' in giocatore:
                    infortuni=giocatore['injuries']
                    for anno, valori in infortuni.items():
                        # Creazione di un dizionario con i valori per il DataFrame
                        riga = { 'Giocatore': giocatore['name'], 'Anno': anno, 'totalDays': valori['totalDays']['value'], 'totalInjuryRate': valori['totalInjuryRate']['value'], 'totalGamesMissed': valori['totalGamesMissed']['value']                    }
                        # Aggiunta della riga alla lista
                        storico_infortuni.loc[len(storico_infortuni)]=riga

                    performances=giocatore['performances']
                    for anno, valori in performances.items():
                        riga = {
                            'Giocatore': giocatore['name'],
                            'Anno': anno,
                            'Competition': valori['competition'],
                            'Club': valori['club'],
                            'Appearances': valori['appearances']['value'],
                            'Goals': valori['goals']['value'],
                            'Assists': valori['assists']['value'],
                            'GoalParticipationPercentage': valori['goalParticipationPercentage']['value'] if valori['goalParticipationPercentage']['value'] is not None else np.nan,
                            'SubstitutionsOn': valori['substitutionsOn']['value'],
                            'SubstitutionsOff': valori['substitutionsOff']['value'],
                            'YellowCards': valori['yellowCards']['value'],
                            'RedCards': valori['redCards']['value'],
                            'MinutesPlayed': valori['minutesPlayed']['value'],
                            'MinutesPerGoal': valori['minutesPerGoal']['value'] if valori['minutesPerGoal']['value'] is not None else np.nan
                        }
                        storico_performances.loc[len(storico_performances)] = riga
                    print("aggiornati infortuni e performance")
                id_array=[giocatore['name'],i['idFantacalcio'],giocatore['id_transfermarket'],giocatore['id_fantaalgoritmo']]
                tabella.loc[len(tabella)]=nuovo_rigo
                #print(f"\n===============================\nvalore nuovo_rigo {nuovo_rigo}")
                id_collection.loc[len(id_collection)]=id_array
                #print(f"\n===============================\nvalore id_array {id_array}")
                try:
                    grafici_valore_mercato(giocatore['name'],giocatore['market_values'],giocatore['name']+'.png')
                    print("----Grafico creato")
                except:
                    print(f"fallita creazione grafico di {name} ")
                    continue
        except:
            print(f"[GRAVE]fallita chiamata di {name}")
            continue
    print('fine-scraping')
    
    with pd.ExcelWriter('scraper_fantaculo_'+datetime.today().strftime('%Y-%m-%d')+'.xlsx') as writer:
        tabella.to_excel(writer, sheet_name='dati')
        df_lista_infortuni.to_excel(writer,sheet_name='infortuni')
        id_collection.to_excel(writer,sheet_name='ids')
        storico_performances.to_excel(writer,sheet_name='performances')
        storico_infortuni.to_excel(writer,sheet_name='storico_infortuni')
        print('file scritto')

def rimuovi_accenti(input_str):
    nfkd_form = unicodedata.normalize('NFD', input_str)
    only_ascii = nfkd_form.encode('ASCII', 'ignore').decode('utf-8')
    return only_ascii

# Funzione ricorsiva per applicare la rimozione accenti a tutte le stringhe in un dizionario o lista
def applica_rimozione_accenti(data):
    if isinstance(data, dict):
        return {k: applica_rimozione_accenti(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [applica_rimozione_accenti(i) for i in data]
    elif isinstance(data, str):
        return rimuovi_accenti(data)
    else:
        return data

def aggiorna_colonna_AI(file_path):
    # Apri il file Excel
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active    
    # Scorri tutte le righe della colonna AI (colonna 35)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=35, max_col=35):
        cell = row[0]
        #print(cell.value)
        if cell.value is None:  # Se la cella è vuota
            # Prendi i valori delle colonne C (3) e B (2) della stessa riga
            team = sheet.cell(row=cell.row, column=3).value
            name = sheet.cell(row=cell.row, column=2).value
            player_url=f'https://fantaculo.it/leghe-srv/api/v1/aste/player?credits=350&flagNoGoal=false&flagModDefense=true&competitionParticipants=12&name={name}&flagSvincolati=false&flagMantra=false&auctionType=pma'
            x = requests.get(player_url)
            giocatore=x.json()
            id_ = giocatore['features']['idFantacalcio']
            # Chiamata alla funzione fantacalcio_calciatore
            nuovo_valore = fantacalcio_calciatore(team, name, id_)
            print(f"Nuova descrizione per {name}:\n{nuovo_valore['descrizione']}")
            # Aggiorna la cella con il valore ottenuto
            sheet.cell(row=cell.row, column=35).value = nuovo_valore['descrizione']
    # Salva il file
    workbook.save(file_path)

def aggiorna_rigo(file_path, mapping_id):
    # Apri il file Excel e carica i diversi DataFrame
    xls = pd.ExcelFile(file_path)
    df = pd.read_excel(xls, sheet_name="dati")
    df_infortuni = pd.read_excel(xls, sheet_name="storico_infortuni")
    df_performances = pd.read_excel(xls, sheet_name="performances")

    listone = aggiorna_dati_fantaculo()
    
    for i in listone:
        if i['name'] not in df['Nome Giocatore'].values and i['name'] != "ILIC" and i['name'] != "CAMARDA" and i['name'] != "ANDERSEN M.K.":
            name = i["name"]
            print(f"Non trovato {name}")
            id = i['idFantacalcio']
            giocatore = get_by_name(name, id, mapping_id)
            #if "injuries" not in giocatore:
             #   print(f"---Skippo figlio di puttana di {name}")
              #  continue
            print(f"Ma {name} che posizione ha {giocatore['Posizione']}")
            # Aggiungi una nuova riga nella tabella 'dati'
            nuovo_rigo = [
                "xyz", giocatore['name'], giocatore['team'], giocatore['role'],
                giocatore['features']['slot'], giocatore['features']['pma'], giocatore['features']['pfc'],
                (int(giocatore['FVM']['Classic'])/1000)*350, giocatore['features']['expectedFantamedia'],
                giocatore['features']['expectedTitolarita'], giocatore['features']['penaltyProbability'],
                giocatore['features']['freeKickProbability'], giocatore['features']['pmaRange'],
                giocatore['features']['pfcRange'], giocatore['features']['lastThreeYearVotoBase'],
                giocatore['features']['lastThreeYearFantamedia'], giocatore['features']['lastThreeYearTitolarity'],
                giocatore['features']['lastFiveYearVotoBase'], giocatore['features']['lastFiveYearFantamedia'],
                giocatore['features']['lastFiveYearTitolarity'], giocatore['features']['lastYearVotoBase'],
                giocatore['features']['lastYearFantamedia'], giocatore['features']['lastYearTitolarity'],
                giocatore['features']['currentSeasonVotoBase'], giocatore['features']['currentSeasonFantamedia'],
                giocatore['features']['currentSeasonTitolarity'], giocatore['Piede'], giocatore['Altezza'],
                giocatore['city'], giocatore['Nato il'], giocatore['Nazionalità'], giocatore['current'],
                giocatore['highest'], giocatore['age'], giocatore['descrizione'], giocatore['fine_contratto'],
                giocatore['Posizione']
            ]
            df.loc[len(df)] = nuovo_rigo
            if 'injuries' in giocatore and  'performances' in giocatore:
                # Aggiungi una nuova riga nella tabella 'infortuni'
                infortuni = giocatore['injuries']
                for anno, valori in infortuni.items():
                    riga_infortuni = [
                        'xyz',
                        giocatore['name'], 
                        anno, valori['totalDays']['value'], 
                        valori['totalInjuryRate']['value'], 
                        valori['totalGamesMissed']['value']
                    ]
                    df_infortuni.loc[len(df_infortuni)] = riga_infortuni
                
                # Aggiungi una nuova riga nella tabella 'performances'
                performances = giocatore['performances']
                for anno, valori in performances.items():
                    riga_performances = [
                        'xyz', giocatore['name'], anno, valori['competition'], valori['club'],
                        valori['appearances']['value'], valori['goals']['value'], valori['assists']['value'],
                        valori['goalParticipationPercentage']['value'] if valori['goalParticipationPercentage']['value'] is not None else np.nan,
                        valori['substitutionsOn']['value'], valori['substitutionsOff']['value'], 
                        valori['yellowCards']['value'], valori['redCards']['value'], 
                        valori['minutesPlayed']['value'], 
                        valori['minutesPerGoal']['value'] if valori['minutesPerGoal']['value'] is not None else np.nan
                    ]
                    df_performances.loc[len(df_performances)] = riga_performances
                
            # Salva i DataFrame aggiornati nel file Excel
            with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                df.to_excel(writer, index=False, sheet_name="dati")    
                df_infortuni.to_excel(writer, index=False, sheet_name="storico_infortuni")
                df_performances.to_excel(writer, index=False, sheet_name="performances")            
                print(f"Aggiornato {name} nelle tabelle dati, infortuni e performances")           

def transfermarket_teams_list():
    headers = { "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"}
    teams = {}
    for team_name, team_list in transfermarket_teams.items():
            url=f'https://www.transfermarkt.it/quickselect/players/{team_list}'
            id_list= requests.get(url, headers=headers)
            teams[team_name]=applica_rimozione_accenti(id_list.json())
    return teams

def fantacalcio_calciatore(team,name,id):
    fc_datas={}
    fantacalcio_calciatore_profile_url=f'https://www.fantacalcio.it/serie-a/squadre/{team}/{name}/{id}'
    headers = { "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"}
    html = requests.get(fantacalcio_calciatore_profile_url, headers=headers)
    soup = BeautifulSoup(html.content, "html.parser")
    section = soup.find('section', {'id': 'player-description'})
    fc_datas['descrizione'] = ''
    for class_name in ['p1', 'li1', 'ul1','bullist','p2']:
        p1_tags = section.find_all(class_=class_name)
        for p1 in p1_tags:
            fc_datas['descrizione']= str(fc_datas['descrizione'])+"\n"+p1.get_text(strip=True)
    dt_tag = soup.find('dt', string='Piede')

    # Trovare il successivo elemento <dd> rispetto a <dt>
    dd_tag = dt_tag.find_next('dd')

    # Trovare il tag <span> all'interno del <dd> e prendere il valore dell'attributo title
    span_title = dd_tag.find('span')['title']
    fc_datas['Piede']=span_title    
    altezza = soup.find('dd', itemprop='height').text
    fc_datas['Altezza'] = altezza
    # Nato il
    nato_il = soup.find('dd', class_='birthdate').text
    fc_datas['Nato il'] = nato_il
    # Nazionalità
    nazionalita = soup.find('dd', class_='nationalities').text
    fc_datas['Nazionalità'] = nazionalita
    # Medie
    medie = {}
    medie_elements = soup.select('.player-stats .group:nth-of-type(1) .badge')
    medie['MV'] = medie_elements[0].text
    medie['FM'] = medie_elements[1].text
    fc_datas['Medie'] = medie
    # Quotazione
    quotazione = {}
    quotazione_elements = soup.select('.player-stats .group:nth-of-type(2) .badge')
    quotazione['Classic'] = quotazione_elements[0].text
    quotazione['Mantra'] = quotazione_elements[1].text
    fc_datas['Quotazione'] = quotazione
    # FVM
    fvm = {}
    fvm_elements = soup.select('.player-stats .group:nth-of-type(3) .badge')
    fvm['Classic'] = fvm_elements[0].text
    fvm['Mantra'] = fvm_elements[1].text
    fc_datas['FVM'] = fvm
    return fc_datas

def grafici_valore_mercato(name, market_values, file_path):
    dates = [pd.to_datetime(mv['x'], unit='ms') for mv in market_values]
    # Converti i valori in milioni di euro
    values = [mv['y'] / 1_000_000 for mv in market_values]
    # Crea il grafico
    plt.figure(figsize=(6, 4))
    plt.plot(dates, values, marker='o', linestyle='-', color='b')
    plt.title(f'Valore di Mercato di {name}')
    plt.xlabel('Data')
    plt.ylabel('Valore in Milioni €')
    plt.grid(True)    
    plt.savefig(file_path, format='png')
    plt.close()

def fantaalgoritmo(name,fantaid):
    
    token = "Bearer eyJhbGciOiJSUzI1NiIsInR5cCIgOiAiSldUIiwia2lkIiA6ICItTzNCNEJ0dGdUclFNNldRQnF6S1NBNUJicjdPV281SnhnbUYzclMzVzdvIn0.eyJleHAiOjE3MjUyOTA1MTQsImlhdCI6MTcyNTI4NjkxNCwiYXV0aF90aW1lIjoxNzI1Mjg2OTEzLCJqdGkiOiI5MmQxMjg1ZC03MGI4LTRmYzAtOTFiOS1kNTkwMjU2ZTU2NGEiLCJpc3MiOiJodHRwczovL2FkbWluLmZhbnRhbGdvcml0bW8uaXQvcmVhbG1zL2ZhbnRhbGdvcml0bW8iLCJhdWQiOiJhY2NvdW50Iiwic3ViIjoiNDg1NzBiYzgtMjVjZi00Yzg5LWJlYjEtOWMzOTkxMjgwNWUxIiwidHlwIjoiQmVhcmVyIiwiYXpwIjoiZmFudGFsZ29yaXRtby1hcHAtcmVnaXN0ZXIiLCJub25jZSI6IjUzNTNkMWY4LTEzMTItNGM5NS1hMzA4LTU4NzdiNTY4ZmQ1YyIsInNlc3Npb25fc3RhdGUiOiIyOGQ4OGZmNC01NWYyLTRhYWUtOTljZC03NzIwZmQyYmY3MzEiLCJhY3IiOiIxIiwiYWxsb3dlZC1vcmlnaW5zIjpbImh0dHBzOi8vYXBwLmZhbnRhbGdvcml0bW8uaXQvKiIsImh0dHBzOi8vYXBwLmZhbnRhbGdvcml0bW8uaXQiXSwicmVhbG1fYWNjZXNzIjp7InJvbGVzIjpbImRlZmF1bHQtcm9sZXMtZmFudGFsZ29yaXRtbyIsIm9mZmxpbmVfYWNjZXNzIiwiZnJlZW1pdW0iLCJ1bWFfYXV0aG9yaXphdGlvbiJdfSwicmVzb3VyY2VfYWNjZXNzIjp7ImFjY291bnQiOnsicm9sZXMiOlsibWFuYWdlLWFjY291bnQiLCJtYW5hZ2UtYWNjb3VudC1saW5rcyIsInZpZXctcHJvZmlsZSJdfX0sInNjb3BlIjoib3BlbmlkIGVtYWlsIHByb2ZpbGUiLCJzaWQiOiIyOGQ4OGZmNC01NWYyLTRhYWUtOTljZC03NzIwZmQyYmY3MzEiLCJlbWFpbF92ZXJpZmllZCI6dHJ1ZSwibmFtZSI6IkFsZXNzaW8gTXVydGEiLCJwcmVmZXJyZWRfdXNlcm5hbWUiOiJhbGVzc2lvLm11cnRhQGdtYWlsLmNvbSIsImdpdmVuX25hbWUiOiJBbGVzc2lvIiwiZmFtaWx5X25hbWUiOiJNdXJ0YSIsImVtYWlsIjoiYWxlc3Npby5tdXJ0YUBnbWFpbC5jb20ifQ.nyqx_K-Jb0EEKWLwZw9mPXf88_YSXmHhNUxmb1aW8X2jK5Q40s7v6DAMXbATHmVj5AM71TzruckWaPPCiUOLJyM7bhPX8P1KlRSB7_GqJzuYDtzx7o0LIUIpnnuU8nU0Zb2IEtDNThVEaB2HJ8dAih70fBDchnfGF00kf8nGEqzTdtRpCHEXIOrS0Eqe9qNbZe9JKoh1Uew8ZpHsiTFM2T3pl5sI4G1D4LzmbmYxgsOmJHBuVxAto3kMc44Fypaia0xoHWyq89J2sjbpbZcxNw6_gUvWHcGCnMTL3pJ_GBAAw3dalwcrtA9hgRsfeZRBXqSkp4oAIEKkA6j049HJiw"
    headers = { 
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36",
        "Authorization": token  # Aggiungi il token come header Authorization
    }
    #print("entro")
    print(name)
    if "ederson" in name:
        name = "Éderson"
    if "zambo" in name:
        name = "anguissa"
    if "n'dicka" in name:
        name = "ndicka"
    
    content=0
    while content < 1 :
        html = requests.get(f"https://api.fantalgoritmo.it/players?page=0&size=50&sort=role%2Casc&sort=surname%2Casc&search={name}", headers=headers)
        print(f"---------search ha risposto {html}")
        giocatore=html.json()
        content=len(giocatore['content'])

        name = name[:-1]
    data_fanta={}
    for x in giocatore['content']:
        #print(f'valore dentro content {x}')
        profile=f"https://api.fantalgoritmo.it/players/{x['id']}"
        get_fantaid= requests.get(profile,headers=headers).json()
        fantacalcioid=get_fantaid['fantacalcioId']
        if int(fantacalcioid) == int(fantaid):
            statistics=f"https://api.fantalgoritmo.it/players/{x['id']}/statistics"
            html2 = requests.get(statistics, headers=headers)
            print(f"---------chiamata alle statistiche {html2.status_code}")
            value_html2=html2.json()
            data_fanta.update(value_html2)
            data_fanta['id_fantaalgoritmo']=x['id']
            print(x)
            print(get_fantaid)
            if 'position' in get_fantaid:
                data_fanta['Posizione']=get_fantaid['position']
                print(f"assegnato position con valore {data_fanta['Posizione']}")
            if "dateContractExpiration" in get_fantaid:
                data_fanta['fine_contratto']= get_fantaid['dateContractExpiration']
    if "fine_contratto" not in data_fanta:
        data_fanta['fine_contratto']  = "Non rompere il cazzo" 
    if "Posizione" not in data_fanta:
        data_fanta['Posizione']  = "Non rompere il cazzo"     
    return data_fanta            

import requests

import requests

def sofascorare():
    listone = aggiorna_dati_fantaculo()
    for i in listone:
        name = i['name'].strip()
        print(f"Name: '{name}'")

        params = {
            "q": name,
            "page": 0
        }
        
        sofascore_url = "http://www.sofascore.com/api/v1/search/all"
        print(f"URL: {sofascore_url}")
        print(f"Params: {params}")

        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36",
            "Accept": "application/json",
            "Accept-Encoding": "gzip, deflate, br",  # Indicazione che accettiamo la compressione
            "Accept-Language": "it-IT,it;q=0.9",
            "DNT": "1",
            "Sec-CH-UA": '"Chromium";v="128", "Not;A=Brand";v="24", "Google Chrome";v="128"',
            "Sec-CH-UA-Mobile": "?0",
            "Sec-CH-UA-Platform": '"Windows"',
            "Upgrade-Insecure-Requests": "1",
        }

        try:
            x = requests.get(sofascore_url, headers=headers, params=params)
            print("Status Code:", x.status_code)
            print("Response URL:", x.url)

            # Verifica se la risposta è compressa
            if 'gzip' in x.headers.get('Content-Encoding', ''):
                print("Decoding gzip content")
                from io import BytesIO
                import gzip
                buf = BytesIO(x.content)
                with gzip.GzipFile(fileobj=buf) as f:
                    content = f.read().decode('utf-8')
            elif 'deflate' in x.headers.get('Content-Encoding', ''):
                print("Decoding deflate content")
                import zlib
                content = zlib.decompress(x.content, -zlib.MAX_WBITS).decode('utf-8')
            else:
                content = x.text  # Nessuna decompressione necessaria

            print("Raw Response Text:", content)
            giocatore = x.json()  # Prova a decodificare solo se il contenuto è valido
            print(giocatore)
        except ValueError:
            print("Errore durante la decodifica della risposta JSON. Contenuto non JSON.")
        except Exception as e:
            print(f"Errore durante la richiesta per {name}: {e}")

        break



#https://fantaculo.it/leghe-srv/api/v1/aste/listone?credits=350&flagNoGoal=false&flagModDefense=true&competitionParticipants=12&name=null&flagSvincolati=false&flagMantra=false&auctionType=pma
if __name__ == "__main__":
    sofascorare()
    #mapping_id=transfermarket_teams_list()
    #aggiorna_rigo("scraper_fantaculo_2024-09-01.xlsx",mapping_id)
    #lista_infortuni=transfermarket_infortuni()
    #scraper(mapping_id,lista_infortuni)
    #aggiorna_rigo("scraper_fantaculo_2024-08-31.xlsx",mapping_id)